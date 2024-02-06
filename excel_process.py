import os
import tkinter as tk
from tkinter import simpledialog, messagebox, Scrollbar
import openpyxl
import datetime
from win32com.client import DispatchEx

cwd = os.getcwd()

class ExcelProcessor:
    def __init__(self, raw_file, calculation_file, summary_file):
        self.raw_file = raw_file
        self.calculation_file = calculation_file
        self.summary_file = summary_file


    def first_1to2(self, date):
        '''
        将1表中的原始数据复制到2表得出计算结果
        '''
        # 读取原始数据表格
        wb1 = openpyxl.load_workbook(self.raw_file,data_only=True, read_only=True)
        # 读取计算表格
        wb2 = openpyxl.load_workbook(self.calculation_file)

        # 获取原始数据sheet
        ws1 = wb1[date]
        # 获取计算表格sheet
        if wb2.__contains__(date) is False:
            self.new_sheet(date)
            wb2 = openpyxl.load_workbook(self.calculation_file)
            print('new_sheet',date)

        ws2 = wb2[date]

        # 复制原始数据到计算表格
        for row in range(4, 29):
            for col in range(4, 8):
                ws2.cell(row=row, column=col - 2).value = ws1.cell(row=row, column=col).value
                print(row, col, ws1.cell(row=row, column=col).value)

        # 复制J4:K28到2.xlsx的sheet`1.29`中的G4:H28
        for row in range(4, 29):
            for col in range(10, 12):
                ws2.cell(row=row, column=col - 3).value = ws1.cell(row=row, column=col).value

        # 补全G4:G28中的空值为1.0，H4:H28中的空值为0.5
        for row in range(4, 29):
            if ws2.cell(row=row, column=7).value is None:
                ws2.cell(row=row, column=7).value = 1.0
            if ws2.cell(row=row, column=8).value is None:
                ws2.cell(row=row, column=8).value = 0.5

        ws2.cell(1,1).value = date

        # 保存计算表格
        wb2.save(self.calculation_file)

        just_open(self.calculation_file)

    def second_2to3(self, date):
        '''将2表中的结果粘贴到3表'''

        # 读取计算表格
        wb2 = openpyxl.load_workbook(self.calculation_file, data_only=True, read_only=True)
        # 读取汇总表格
        wb3 = openpyxl.load_workbook(self.summary_file)

        # 获取计算表格sheet
        ws2 = wb2[date]
        # 获取汇总表格sheet
        ws3 = wb3['Sheet1']

        print(ws3.cell(3,1).value)

        # 获取计算结果数据
        result_data = []
        for col in range(11, 15):
            result_data.append(ws2.cell(row=29, column=col).value)
        for col in range(7, 9):
            result_data.append(ws2.cell(row=29, column=col).value)

        excel_date = convert_to_excel_date(date)

        # 获取日期对应的行
        date_column = ws3['A']
        for idx, cell in enumerate(date_column):
            if cell.value == excel_date:
                row_index = idx + 1  # 因为索引从0开始，所以需要+1
                break
        else:
            row_index = len(date_column) + 1
            ws3.cell(row=row_index, column=1).value = excel_date

        # 将计算结果数据复制到汇总表格中
        for col, value in zip(range(11, 17), result_data):
            print(value)
            ws3.cell(row=row_index, column=col).value = value

        # 保存汇总表格
        wb3.save(self.summary_file)

    def new_sheet(self, date):
        # 读取计算表格
        wb2 = openpyxl.load_workbook(self.calculation_file)

        # 获取计算表格sheet
        ws2 = wb2['模板']
        
        new_ws2 = wb2.copy_worksheet(ws2)
        new_ws2.title = date
        wb2.save(self.calculation_file)

def just_open(filename):

    xlApp = DispatchEx("Excel.Application")

    xlApp.Visible = False

    xlBook = xlApp.Workbooks.Open(os.path.join(cwd, filename))

    xlBook.Save()

    xlBook.Close()


def convert_to_excel_date(date_string):
    # 将日期字符串解析为 datetime 对象
    date_obj = datetime.datetime.strptime(date_string, '%y.%m.%d')

    # 计算日期与 1899 年 12 月 31 日之间的天数差
    delta_days = (date_obj - datetime.datetime(1899, 12, 31)).days

    # 将天数差加上 Excel 日期的起始值
    excel_date = delta_days + 1  # Excel日期起始值是1900年1月1日

    return excel_date


# processor = ExcelProcessor('1.xlsx', '2.xlsx', '3.xlsx')
# date = '24.2.4'
# processor.first_1to2(date)
# processor.second_2to3(date)


class ExcelProcessorGUI:
    def __init__(self, root, raw_file, calculation_file, summary_file):
        self.root = root
        self.root.title("Excel Processor")

        # 创建一个自定义字体
        custom_font = ("Arial", 14)

        # 创建一个滚动条
        self.scrollbar = Scrollbar(root, orient=tk.VERTICAL)

        # 创建列表框并关联滚动条
        self.sheet_listbox = tk.Listbox(root, selectmode=tk.MULTIPLE, font=custom_font, height=25,
                                         yscrollcommand=self.scrollbar.set)

        # 设置滚动条与列表框的关联
        self.scrollbar.config(command=self.sheet_listbox.yview)

        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.sheet_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)


        self.run_button = tk.Button(root, text="运行", command=self.run_process, font=custom_font)  # 设置按钮的字体
        self.run_button.pack()

        self.error_file = "error_sheets.txt"
        self.error_sheets = []

        self.processor = ExcelProcessor(raw_file, calculation_file, summary_file)

        self.populate_sheet_listbox()

    def populate_sheet_listbox(self):
        try:
            wb = openpyxl.load_workbook(self.processor.raw_file, read_only=True)
            sheet_names = wb.sheetnames
            for sheet_name in sheet_names:
                self.sheet_listbox.insert(0, sheet_name)  # 在索引 0 处插入新的 sheet 名称
        except Exception as e:
            messagebox.showerror("错误", f"无法加载工作簿: {e}")

    def run_process(self):
        selected_sheets = self.sheet_listbox.curselection()
        if not selected_sheets:
            messagebox.showwarning("警告", "请至少选择一个表格进行处理")
            return

        for index in selected_sheets:
            sheet_name = self.sheet_listbox.get(index)
            if not self.is_valid_date(sheet_name):
                messagebox.showinfo("提示", f"表格 {sheet_name} 的名称不符合 %y.%m.%d 格式，请手动输入一个日期")
                new_date = self.get_valid_date_from_user()
                if new_date:
                    self.rename_sheet(sheet_name, new_date)
                    sheet_name = new_date
                else:
                    continue

            try:
                self.processor.first_1to2(sheet_name)
                self.processor.second_2to3(sheet_name)
                messagebox.showinfo("完成", f"{sheet_name} 处理完成")
            except Exception as e:
                messagebox.showerror("错误", f"处理 {sheet_name} 时出现错误: {e}")
                self.error_sheets.append(sheet_name)
        if len(self.error_sheets)>=1:
            self.save_error_sheets()

    def is_valid_date(self, date_string):
        try:
            datetime.datetime.strptime(date_string, '%y.%m.%d')
            return True
        except ValueError:
            return False

    def get_valid_date_from_user(self):
        new_date = simpledialog.askstring("输入日期", "请输入一个符合 %y.%m.%d 格式的日期：")
        return new_date.strip() if new_date else None

    def rename_sheet(self, old_name, new_name):
        try:
            wb1 = openpyxl.load_workbook(self.processor.raw_file)
            ws = wb1[old_name]
            ws.title = new_name
            wb1.save(self.processor.raw_file)

        except Exception as e:
            messagebox.showerror("错误", f"重命名 {old_name} 时出现错误: {e}")

    def save_error_sheets(self):
        try:
            with open(self.error_file, 'w') as file:
                for sheet_name in self.error_sheets:
                    file.write(sheet_name + '\n')
            messagebox.showinfo("完成", f"出错的表格名称已保存到文件: {self.error_file}")
        except Exception as e:
            messagebox.showerror("错误", f"保存出错的表格名称时出现错误: {e}")

root = tk.Tk()
app = ExcelProcessorGUI(root, '1原油流量计记录.xlsx', '2.xlsx', '3.xlsx')
root.mainloop()
